#!/usr/bin/env python3
from __future__ import annotations

from pathlib import Path
from typing import List

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from analysis.output_paths import SCOPES

try:
    from openpyxl.drawing.image import Image as XLImage
except Exception:  # pragma: no cover
    XLImage = None

OUTPUT_ROOT = Path("output")


def read_csv_or_note(path: Path) -> pd.DataFrame:
    if not path.exists():
        return pd.DataFrame([{"note": f"Missing file: {path.as_posix()}"}])
    try:
        return pd.read_csv(path)
    except Exception as e:
        return pd.DataFrame([{"note": f"Could not read {path.as_posix()}: {e}"}])


def safe_cell_value(value: object) -> object:
    try:
        if pd.isna(value):
            return None
    except Exception:
        pass
    if isinstance(value, (list, tuple, set, dict)):
        return str(value)
    return value


def write_dataframe(ws, df: pd.DataFrame, start_row: int = 1, start_col: int = 1) -> int:
    if df.empty:
        ws.cell(row=start_row, column=start_col, value="(no rows)")
        return start_row + 1

    cols = [str(c) for c in df.columns]
    for j, col in enumerate(cols, start=start_col):
        cell = ws.cell(row=start_row, column=j, value=col)
        cell.font = Font(name="Calibri", size=10, bold=True)

    r = start_row + 1
    for row in df.itertuples(index=False, name=None):
        for j, val in enumerate(row, start=start_col):
            ws.cell(row=r, column=j, value=safe_cell_value(val))
        r += 1
    return r


def embed_image(ws, image_path: Path, anchor: str, max_w: float = 430.0, max_h: float = 220.0) -> None:
    if XLImage is None or not image_path.exists():
        return
    try:
        img = XLImage(str(image_path))
        w = float(getattr(img, "width", 1.0))
        h = float(getattr(img, "height", 1.0))
        if w > 0 and h > 0:
            scale = min(max_w / w, max_h / h, 1.0)
            img.width = int(w * scale)
            img.height = int(h * scale)
        ws.add_image(img, anchor)
    except Exception:
        return


def collect_scope_figures(scope_dir: Path) -> List[Path]:
    return sorted([p for p in scope_dir.rglob("*.png") if p.is_file()])


def add_page_one_figures(ws, scope: str, scope_dir: Path) -> None:
    ws.title = "Page1_Figures"
    ws["A1"] = f"{scope.upper()} Study Figures"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True)

    figures = collect_scope_figures(scope_dir)
    if not figures:
        ws["A3"] = "No figures found."
        return

    for idx, fig_path in enumerate(figures):
        col = "A" if idx % 2 == 0 else "J"
        row = 3 + (idx // 2) * 20
        embed_image(ws, fig_path, f"{col}{row}")


def add_page_two_raw_long(ws, raw_long: pd.DataFrame) -> None:
    ws.title = "Page2_Raw_Long"
    ws["A1"] = "Raw Data Long Dataset"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True)
    write_dataframe(ws, raw_long, start_row=3, start_col=1)
    ws.freeze_panes = "A4"


def add_page_three_pca(
    ws,
    scope_dir: Path,
    pca_var: pd.DataFrame,
    pca_scores: pd.DataFrame,
    pca_loadings: pd.DataFrame,
) -> None:
    ws.title = "Page3_PCA"
    ws["A1"] = "PCA Findings"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True)

    row = 3
    ws.cell(row=row, column=1, value="Explained Variance").font = Font(name="Calibri", size=10, bold=True)
    row = write_dataframe(ws, pca_var, start_row=row + 1, start_col=1) + 1

    ws.cell(row=row, column=1, value="PCA Scores").font = Font(name="Calibri", size=10, bold=True)
    row = write_dataframe(ws, pca_scores, start_row=row + 1, start_col=1) + 1

    ws.cell(row=row, column=1, value="PCA Loadings").font = Font(name="Calibri", size=10, bold=True)
    write_dataframe(ws, pca_loadings, start_row=row + 1, start_col=1)

    embed_image(ws, scope_dir / "pca" / "Fig1.png", "J3", max_w=430, max_h=240)
    ws.freeze_panes = "A4"


def build_scope_workbook(scope: str) -> Path:
    scope_dir = OUTPUT_ROOT / scope
    if not scope_dir.exists():
        raise FileNotFoundError(f"Missing scope directory: {scope_dir}")

    raw_long = read_csv_or_note(scope_dir / "metadata" / "spectral" / "spectra_long.csv")
    pca_var = read_csv_or_note(scope_dir / "metadata" / "pca" / "pca_explained_variance.csv")
    pca_scores = read_csv_or_note(scope_dir / "metadata" / "pca" / "pca_scores.csv")
    pca_loadings = read_csv_or_note(scope_dir / "metadata" / "pca" / "pca_loadings.csv")

    wb = Workbook()
    add_page_one_figures(wb.active, scope, scope_dir)
    add_page_two_raw_long(wb.create_sheet(), raw_long)
    add_page_three_pca(wb.create_sheet(), scope_dir, pca_var, pca_scores, pca_loadings)

    workbook_path = scope_dir / f"{scope}_executive_report.xlsx"
    fallback = scope_dir / f"{scope}_executive_report_new.xlsx"
    try:
        wb.save(workbook_path)
        if fallback.exists():
            try:
                fallback.unlink()
            except OSError:
                pass
        return workbook_path
    except PermissionError:
        if workbook_path.exists():
            return workbook_path
        wb.save(fallback)
        return fallback


def main() -> int:
    generated: List[Path] = []
    failures: List[str] = []
    for scope in SCOPES:
        try:
            generated.append(build_scope_workbook(scope))
        except Exception as e:
            failures.append(f"{scope}: {e}")

    if generated:
        print("Wrote executive workbooks:")
        for path in generated:
            print(f"  {path}")
    if failures:
        print("Executive workbook generation failures:")
        for row in failures:
            print(f"  {row}")
        return 2
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
